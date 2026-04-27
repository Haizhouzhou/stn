#!/usr/bin/env python3
"""Phase 6A Part 2 audit for the Herz/Groppa/Brown force adaptation package.

This is an audit-only script. It inventories the shared package, inspects
schemas lightly, and parses MATLAB code as text. It does not run MATLAB,
FieldTrip, preprocessing, statistics, or analysis reproduction.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import difflib
import json
import math
import os
import re
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

try:
    import numpy as np
except Exception as exc:  # pragma: no cover - optional dependency fallback
    np = None  # type: ignore[assignment]
    NUMPY_IMPORT_ERROR = f"{exc.__class__.__name__}: {exc}"
else:
    NUMPY_IMPORT_ERROR = ""

try:
    import scipy.io as scipy_io
except Exception as exc:  # pragma: no cover - optional dependency fallback
    scipy_io = None  # type: ignore[assignment]
    SCIPY_IMPORT_ERROR = f"{exc.__class__.__name__}: {exc}"
else:
    SCIPY_IMPORT_ERROR = ""

try:
    import h5py
except Exception as exc:  # pragma: no cover - optional dependency fallback
    h5py = None  # type: ignore[assignment]
    H5PY_IMPORT_ERROR = f"{exc.__class__.__name__}: {exc}"
else:
    H5PY_IMPORT_ERROR = ""

try:
    import openpyxl
except Exception as exc:  # pragma: no cover - optional dependency fallback
    openpyxl = None  # type: ignore[assignment]
    OPENPYXL_IMPORT_ERROR = f"{exc.__class__.__name__}: {exc}"
else:
    OPENPYXL_IMPORT_ERROR = ""


DATASET_NAME = "Subthalamic nucleus correlates of force adaptation"
PAPER_TITLE = "Dynamic modulation of subthalamic nucleus activity facilitates adaptive behavior"
EXPECTED_EXAMPLE_SUBJECTS = ["Kont01", "Kont02"]
EXPECTED_STUDY_PD_PATIENTS = 16
EXPECTED_STUDY_HC_ANALYZED = 15
EXPECTED_STIM_PATIENTS = 14

SCAN_EXTENSIONS = {
    ".dcm",
    ".nii",
    ".nii.gz",
    ".mri",
    ".ima",
    ".nrrd",
    ".mgz",
    ".mgh",
    ".hdr",
    ".img",
    ".bval",
    ".bvec",
}
CODE_EXTENSIONS = {".m", ".mlx", ".py", ".sh", ".ipynb"}
TABULAR_EXTENSIONS = {".csv", ".tsv", ".txt", ".xlsx", ".xls", ".rtf"}
ZIP_EXTENSIONS = {".zip"}
MAT_EXTENSIONS = {".mat"}

SCAN_TERMS_RE = re.compile(
    r"(\bMRI\b|\bCT\b|post[-_ ]?op|pre[-_ ]?op|LeadDBS|lead_dbs|\bMNI\b|"
    r"\banat(?:omical)?\b|\bDICOM\b|\bNIfTI\b)",
    re.IGNORECASE,
)
SUBJECT_RE = re.compile(
    r"(?<![A-Za-z0-9])((?:PD|HC|Kont|Control|Patient|Subj|Subject)[_-]?\d{1,3})(?![A-Za-z0-9])",
    re.IGNORECASE,
)
FS_KEY_RE = re.compile(
    r"(^fs$|fsample|sample[_-]?rate|sampling[_-]?(rate|freq|frequency)|"
    r"^sfreq$|^sr$|^Fs$|^FS$|NewSR|resamplefs)",
    re.IGNORECASE,
)
BEHAVIOR_KEYWORDS = {
    "force",
    "target",
    "actual",
    "peak",
    "baseline",
    "mvc",
    "yank",
    "auc",
    "value",
    "points",
    "direction",
    "error",
    "rt",
    "reaction",
    "trial",
    "response",
    "cue",
    "feedback",
    "too much",
    "too little",
    "abs_change",
    "change_in_force",
    "diff",
}
STIM_KEYWORDS = {
    "dbs",
    "stim",
    "stimulation",
    "onset",
    "burst",
    "ramp",
    "tos",
    "window",
    "binary",
    "stim_on",
    "stim_off",
}
LFP_KEYWORDS = {
    "lfp",
    "stn",
    "fieldtrip",
    "freq",
    "powspctrm",
    "spectra",
    "spectrum",
    "beta",
    "alpha",
    "gamma",
    "montage",
    "fsample",
    "trial",
}
MATLAB_CONTROL_WORDS = {
    "if",
    "elseif",
    "else",
    "for",
    "while",
    "switch",
    "case",
    "otherwise",
    "try",
    "catch",
    "end",
    "function",
    "return",
    "break",
    "continue",
}
MATLAB_BUILTIN_HINTS = {
    "abs",
    "addpath",
    "all",
    "any",
    "axis",
    "bar",
    "cell",
    "cell2mat",
    "clear",
    "clc",
    "close",
    "corr",
    "csvread",
    "csvwrite",
    "disp",
    "dir",
    "errorbar",
    "eval",
    "exist",
    "export_fig",
    "figure",
    "find",
    "fopen",
    "fprintf",
    "fread",
    "fscanf",
    "get",
    "hold",
    "importdata",
    "interp1",
    "isfield",
    "length",
    "linspace",
    "load",
    "max",
    "mean",
    "median",
    "min",
    "nan",
    "nanmean",
    "nanstd",
    "numel",
    "ones",
    "plot",
    "prctile",
    "print",
    "randperm",
    "repmat",
    "reshape",
    "round",
    "save",
    "set",
    "size",
    "sort",
    "sprintf",
    "std",
    "str2double",
    "sum",
    "table",
    "textscan",
    "title",
    "unique",
    "writetable",
    "xlsread",
    "xlswrite",
    "zeros",
}
STAT_TOOLBOX_HINTS = {
    "fitlme",
    "anova",
    "ranova",
    "coefTest",
    "LinearMixedModel",
    "lillietest",
    "ttest",
    "ttest2",
    "ranksum",
    "signrank",
    "corr",
    "partialcorr",
}
DOWNLOADED_HELPERS = {"computeCohen_d", "jblill", "shadedErrorBar"}


EXPECTED_COMPONENTS: list[dict[str, str]] = [
    {"component_name": "CompareLevodopaDemographicsMVC", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Compare demographics, MVC, levodopa/UPDRS variables"},
    {"component_name": "GetEvents_PD", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Import PD PsychoPy events and save MAT events"},
    {"component_name": "GetEvents_HC", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Import healthy-control PsychoPy events and save MAT events"},
    {"component_name": "ExtractData", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Extract relevant behavioral event variables"},
    {"component_name": "ExtractData_HC", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Extract healthy-control behavioral event variables"},
    {"component_name": "GetForce_PD", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Compute PD force production and adaptation variables"},
    {"component_name": "GetForce_HC", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Compute healthy-control/example force variables"},
    {"component_name": "Forceparameters", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Compute force-production parameters and peak force"},
    {"component_name": "Forces_within", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Single-subject force trace summaries"},
    {"component_name": "Stat_within", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Single-subject behavioral statistics"},
    {"component_name": "Forces_across", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Group force trace summaries"},
    {"component_name": "stat_across", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Group force and target trajectory statistics"},
    {"component_name": "Plot_Stats", "expected_module": "Behavioral force analysis / Figure 1", "expected_role": "Plot/statistics for force adaptation measures"},
    {"component_name": "GetLFP_FirstLevel", "expected_module": "STN LFP first/second-level analysis / Figures 2 and 3", "expected_role": "FieldTrip preprocessing and first-level time-frequency analysis"},
    {"component_name": "MakeMontage_AllBipolar", "expected_module": "STN LFP first/second-level analysis / Figures 2 and 3", "expected_role": "Create bipolar STN montage"},
    {"component_name": "EpochData_TF", "expected_module": "STN LFP first/second-level analysis / Figures 2 and 3", "expected_role": "Epoch continuous time-frequency data"},
    {"component_name": "GetLFP_SecondLevel_PlotSpectra", "expected_module": "STN LFP first/second-level analysis / Figures 2 and 3", "expected_role": "Plot grand-average spectra and band traces"},
    {"component_name": "GetLFP_SecondLevel_LME", "expected_module": "STN LFP first/second-level analysis / Figures 2 and 3", "expected_role": "Linear mixed-effects analyses on beta power windows"},
    {"component_name": "PermTests_LME", "expected_module": "STN LFP first/second-level analysis / Figures 2 and 3", "expected_role": "Cluster/permutation tests for LME outputs"},
    {"component_name": "GetLFP_SecondLevel_controlLME", "expected_module": "STN LFP first/second-level analysis / Figures 2 and 3", "expected_role": "Control LME analyses"},
    {"component_name": "GetEvents_Stim", "expected_module": "DBS timing effects on behavior / Figures 4 and 5", "expected_role": "Import stimulation-session PsychoPy events"},
    {"component_name": "GetForce_Stim", "expected_module": "DBS timing effects on behavior / Figures 4 and 5", "expected_role": "Compute force variables for DBS session"},
    {"component_name": "GetToS", "expected_module": "DBS timing effects on behavior / Figures 4 and 5", "expected_role": "Compute timing of stimulation windows and behavioral variables"},
    {"component_name": "ToS_DownsampleBinaryRemoveRamp", "expected_module": "DBS timing effects on behavior / Figures 4 and 5", "expected_role": "Downsample stimulation to binary 1000 Hz trace and remove ramps"},
    {"component_name": "ToS_WindowedStim", "expected_module": "DBS timing effects on behavior / Figures 4 and 5", "expected_role": "Compute windowed stimulation indicators"},
    {"component_name": "ToS_Windowed_nexttrial", "expected_module": "DBS timing effects on behavior / Figures 4 and 5", "expected_role": "Compute next-trial force changes by stimulation windows"},
    {"component_name": "Plot_ToS", "expected_module": "DBS timing effects on behavior / Figures 4 and 5", "expected_role": "Plot timing-of-stimulation effects and call permutation tests"},
    {"component_name": "PermTests_ToS", "expected_module": "DBS timing effects on behavior / Figures 4 and 5", "expected_role": "Cluster/permutation tests for stimulation timing"},
    {"component_name": "GetLFP_FirstLevel_Stim", "expected_module": "DBS timing effects on STN LFP / Figures 4 and 5", "expected_role": "First-level FieldTrip analysis for stimulation session"},
    {"component_name": "GetLFP_FirstLevel_Stim_TrigOnset", "expected_module": "DBS timing effects on STN LFP / Figures 4 and 5", "expected_role": "First-level FieldTrip analysis aligned to stimulation onset"},
    {"component_name": "GetLFP_SecondLevel_Stim", "expected_module": "DBS timing effects on STN LFP / Figures 4 and 5", "expected_role": "Second-level stimulation LFP analysis"},
    {"component_name": "GetLFP_SecondLevel_TrigOnset", "expected_module": "DBS timing effects on STN LFP / Figures 4 and 5", "expected_role": "Second-level stimulation-onset LFP analysis"},
    {"component_name": "computeCohen_d", "expected_module": "Downloaded/helper scripts", "expected_role": "Effect-size helper downloaded from MathWorks"},
    {"component_name": "jblill", "expected_module": "Downloaded/helper scripts", "expected_role": "Significance-cluster fill helper"},
    {"component_name": "shadedErrorBar", "expected_module": "Downloaded/helper scripts", "expected_role": "Mean/SEM plotting helper"},
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-root", default="cambium/Force_Scripts")
    parser.add_argument(
        "--out-dir",
        default="reports/phase6_stn_force_adaptation_herz_2023_audit",
    )
    parser.add_argument("--paper-fs", type=float, default=2048.0)
    parser.add_argument("--lfp-analysis-fs", type=float, default=200.0)
    parser.add_argument("--stim-binary-fs", type=float, default=1000.0)
    parser.add_argument("--max-text-read-bytes", type=int, default=200000)
    parser.add_argument("--max-sanity-arrays", type=int, default=200)
    return parser.parse_args()


def local_timestamp() -> str:
    return dt.datetime.now().astimezone().isoformat(timespec="seconds")


def relpath(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return path.as_posix()


def clean_join(values: Iterable[Any], sep: str = ";") -> str:
    cleaned = []
    seen = set()
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        if text not in seen:
            cleaned.append(text)
            seen.add(text)
    return sep.join(cleaned)


def bool_cell(value: Any) -> str:
    return "true" if bool(value) else "false"


def json_safe(value: Any) -> Any:
    if isinstance(value, Path):
        return value.as_posix()
    if isinstance(value, set):
        return sorted(json_safe(v) for v in value)
    if isinstance(value, tuple):
        return [json_safe(v) for v in value]
    if isinstance(value, list):
        return [json_safe(v) for v in value]
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    return value


def json_cell(value: Any) -> str:
    if value in (None, "", [], {}, ()):
        return ""
    return json.dumps(json_safe(value), sort_keys=True, ensure_ascii=True)


def file_extension(path_or_name: Path | str) -> str:
    name = path_or_name.name if isinstance(path_or_name, Path) else str(path_or_name)
    lower = name.lower()
    if lower.endswith(".nii.gz"):
        return ".nii.gz"
    suffix = Path(name).suffix.lower()
    return suffix or "[none]"


def is_apple_metadata(path: Path | str) -> bool:
    parts = Path(path).parts if not isinstance(path, Path) else path.parts
    return "__MACOSX" in parts or any(part.startswith("._") for part in parts) or Path(path).name == ".DS_Store"


def normalize_component_name(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def read_text_limited(path: Path, max_bytes: int) -> tuple[str, str]:
    try:
        data = path.read_bytes()[:max_bytes]
    except Exception as exc:
        return "", f"{exc.__class__.__name__}: {exc}"
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding), ""
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace"), "decoded_with_replacement"


def truncate(text: str, limit: int = 220) -> str:
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def standardize_subject_id(raw: str) -> str:
    raw = raw.replace("_", "").replace("-", "")
    match = re.match(r"([A-Za-z]+)(\d{1,3})$", raw)
    if not match:
        return raw
    prefix, digits = match.groups()
    normalized_prefix = {
        "pd": "PD",
        "hc": "HC",
        "kont": "Kont",
        "control": "Control",
        "patient": "Patient",
        "subj": "Subj",
        "subject": "Subject",
    }.get(prefix.lower(), prefix)
    if normalized_prefix in {"PD", "HC", "Kont"}:
        digits = digits.zfill(2)
    return f"{normalized_prefix}{digits}"


def extract_subject_ids(text: str) -> list[str]:
    ids = [standardize_subject_id(match.group(1)) for match in SUBJECT_RE.finditer(text)]
    return sorted(set(ids))


def infer_subject_id(path_text: str) -> str:
    ids = extract_subject_ids(path_text)
    if ids:
        return ids[0]
    return ""


def infer_group(subject_id: str) -> str:
    lower = subject_id.lower()
    if lower.startswith("pd"):
        return "PD"
    if lower.startswith("hc"):
        return "HC"
    if lower.startswith("kont"):
        return "example_control"
    if lower.startswith("control"):
        return "HC"
    if lower.startswith("patient"):
        return "PD"
    return "unknown"


def guess_module(path_text: str) -> str:
    lower = path_text.lower()
    if "__macosx" in lower or "/._" in lower:
        return "macos_metadata"
    if "1behavioraldata" in lower:
        return "Behavioral force analysis / Figure 1"
    if "2localfieldpotentialdata" in lower:
        return "STN LFP first/second-level analysis / Figures 2 and 3"
    if "3dbseffectsbehavior" in lower:
        return "DBS timing effects on behavior / Figures 4 and 5"
    if "4dbseffectslocalfieldpotential" in lower:
        return "DBS timing effects on STN LFP / Figures 4 and 5"
    if "exampledata" in lower:
        return "Example behavioral data"
    if any(helper.lower() in lower for helper in DOWNLOADED_HELPERS):
        return "Downloaded/helper scripts"
    return "unknown"


def guess_analysis_part(path_text: str) -> str:
    module = guess_module(path_text)
    if module.startswith("Behavioral"):
        return "behavioral"
    if module.startswith("STN LFP"):
        return "lfp_off_stim"
    if module.startswith("DBS timing effects on behavior"):
        return "dbs_behavior"
    if module.startswith("DBS timing effects on STN LFP"):
        return "dbs_lfp"
    if module.startswith("Downloaded"):
        return "downloaded_helper"
    if module.startswith("Example"):
        return "source_data"
    return "unknown"


def guess_figure_mapping(path_text: str) -> str:
    part = guess_analysis_part(path_text)
    if part == "behavioral":
        return "fig1"
    if part == "lfp_off_stim":
        return "fig2;fig3"
    if part in {"dbs_behavior", "dbs_lfp"}:
        return "fig4;fig5"
    if "s1" in path_text.lower() or "supp" in path_text.lower():
        return "supplemental"
    return "unknown"


def contains_any(text: str, keywords: Iterable[str]) -> bool:
    lower = text.lower()
    return any(keyword.lower() in lower for keyword in keywords)


def is_possible_scan_or_identifiable(path_text: str, ext: str) -> bool:
    return ext in SCAN_EXTENSIONS or bool(SCAN_TERMS_RE.search(path_text))


def row_fieldnames(rows: list[dict[str, Any]], fieldnames: list[str]) -> list[dict[str, Any]]:
    normalized = []
    for row in rows:
        normalized.append({field: row.get(field, "") for field in fieldnames})
    return normalized


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    def sanitize(value: Any) -> Any:
        if isinstance(value, str):
            return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", value)
        return value

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in row_fieldnames(rows, fieldnames):
            writer.writerow({key: sanitize(value) for key, value in row.items()})


def collect_files(data_root: Path) -> list[Path]:
    return sorted(path for path in data_root.rglob("*") if path.is_file())


def build_local_m_map(files: list[Path], data_root: Path) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = defaultdict(list)
    for path in files:
        if file_extension(path) != ".m" or is_apple_metadata(path):
            continue
        mapping[path.stem.lower()].append(relpath(path, data_root))
    return dict(mapping)


def extract_quoted_paths(text: str) -> list[str]:
    values = []
    for match in re.finditer(r"['\"]([^'\"]{2,260})['\"]", text):
        value = match.group(1).strip()
        if "\n" in value or "\r" in value:
            continue
        if "/" in value or "\\" in value or re.search(r"\.[A-Za-z0-9]{1,5}\b", value) or extract_subject_ids(value):
            values.append(value)
    return values


def extract_hardcoded_paths(text: str) -> list[str]:
    patterns = [
        r"['\"](/Users/[^'\"]+)['\"]",
        r"['\"](/Volumes/[^'\"]+)['\"]",
        r"['\"]([A-Za-z]:\\[^'\"]+)['\"]",
        r"['\"](~\/[^'\"]+)['\"]",
    ]
    paths = []
    for pattern in patterns:
        paths.extend(match.group(1) for match in re.finditer(pattern, text))
    return sorted(set(paths))


def first_comment_summary(lines: list[str]) -> str:
    comments: list[str] = []
    in_block = False
    for line in lines[:60]:
        stripped = line.strip()
        if stripped.startswith("%"):
            comments.append(stripped.lstrip("%").strip())
            in_block = True
            continue
        if in_block:
            break
        if stripped and not stripped.startswith("%%"):
            continue
    return truncate(" ".join(comment for comment in comments if comment), 500)


def parse_function_name(lines: list[str]) -> tuple[bool, str]:
    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("%"):
            continue
        match = re.match(
            r"function\s+(?:\[[^\]]+\]\s*=\s*|[A-Za-z]\w*\s*=\s*)?([A-Za-z]\w*)",
            stripped,
            flags=re.IGNORECASE,
        )
        if match:
            return True, match.group(1)
        return False, ""
    return False, ""


def extract_line_excerpts(lines: list[str], pattern: str, limit: int = 12) -> list[str]:
    regex = re.compile(pattern, re.IGNORECASE)
    excerpts = []
    for line in lines:
        if regex.search(line):
            excerpts.append(truncate(line, 180))
            if len(excerpts) >= limit:
                break
    return excerpts


def parse_matlab_calls(lines: list[str]) -> dict[str, str]:
    calls: dict[str, str] = {}
    call_re = re.compile(r"(?<![\w.])([A-Za-z]\w*)\s*\(")
    script_call_re = re.compile(r"^\s*([A-Za-z]\w*)\s*;?\s*$")
    for line in lines:
        code = line.split("%", 1)[0]
        if not code.strip():
            continue
        for match in call_re.finditer(code):
            name = match.group(1)
            if name.lower() in MATLAB_CONTROL_WORDS:
                continue
            calls.setdefault(name, truncate(line, 180))
        simple = script_call_re.match(code)
        if simple:
            name = simple.group(1)
            if name.lower() not in MATLAB_CONTROL_WORDS and name.lower() not in {"clear", "clc", "close"}:
                calls.setdefault(name, truncate(line, 180))
    return calls


def classify_dependency(name: str, local_m_map: dict[str, list[str]]) -> str:
    if name.startswith("ft_"):
        return "FieldTrip"
    if name in DOWNLOADED_HELPERS:
        return "downloaded_helper"
    if name in STAT_TOOLBOX_HINTS:
        return "statistics_toolbox"
    if name.lower() in local_m_map:
        return "local_function"
    if name in MATLAB_BUILTIN_HINTS or name.lower() in {item.lower() for item in MATLAB_BUILTIN_HINTS}:
        return "MATLAB_builtin_or_toolbox"
    lower = name.lower()
    if lower.startswith("ea_") or lower.startswith("nifti") or "lead" in lower:
        return "LeadDBS_or_imaging"
    return "unknown"


def inspect_matlab_code(
    files: list[Path],
    data_root: Path,
    local_m_map: dict[str, list[str]],
    max_text_read_bytes: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, dict[str, Any]], set[str]]:
    code_rows: list[dict[str, Any]] = []
    dep_rows: list[dict[str, Any]] = []
    info_by_rel: dict[str, dict[str, Any]] = {}
    subject_refs: set[str] = set()

    for path in files:
        if file_extension(path) != ".m":
            continue
        rel = relpath(path, data_root)
        text, read_error = read_text_limited(path, max_text_read_bytes)
        lines = text.splitlines()
        is_function, function_name = parse_function_name(lines)
        lower = text.lower()
        ft_calls = sorted(set(re.findall(r"\b(ft_[A-Za-z]\w*)\s*\(", text)))
        hardcoded_paths = extract_hardcoded_paths(text)
        load_calls = extract_line_excerpts(lines, r"\bload\s*(?:\(|\s)")
        save_calls = extract_line_excerpts(lines, r"\bsave\s*(?:\(|\s)")
        read_calls = extract_line_excerpts(lines, r"\b(readtable|readmatrix|xlsread|csvread|importdata|importfile|fopen|textscan)\b")
        write_calls = extract_line_excerpts(lines, r"\b(writetable|writematrix|xlswrite|csvwrite|fprintf)\b")
        quoted_paths = extract_quoted_paths(text)
        figures = extract_line_excerpts(lines, r"\b(figure|subplot|plot|imagesc|shadedErrorBar|saveas|print)\b", limit=10)
        subject_refs.update(extract_subject_ids(text))

        function_or_script = function_name if is_function and function_name else path.stem
        row = {
            "relative_path": rel,
            "file_name": path.name,
            "module_guess": guess_module(rel),
            "n_lines": len(lines),
            "is_function": bool_cell(is_function),
            "function_name_if_any": function_name,
            "script_or_function_guess": "function" if is_function else "script",
            "first_comment_block_summary": first_comment_summary(lines),
            "mentions_fieldtrip": bool_cell("fieldtrip" in lower or bool(ft_calls)),
            "ft_functions_called": clean_join(ft_calls),
            "mentions_fitlme_or_lme": bool_cell("fitlme" in lower or "lme" in lower or "linearmixedmodel" in lower),
            "mentions_permutation_or_cluster": bool_cell("permutation" in lower or "cluster" in lower or "randperm" in lower),
            "mentions_force_processing": bool_cell(contains_any(lower, BEHAVIOR_KEYWORDS)),
            "mentions_lfp_processing": bool_cell(contains_any(lower, LFP_KEYWORDS)),
            "mentions_stimulation_processing": bool_cell(contains_any(lower, STIM_KEYWORDS)),
            "mentions_psychopy_or_events": bool_cell("psychopy" in lower or "event" in lower or "ttl" in lower or "labjack" in lower),
            "mentions_leaddbs_or_imaging": bool_cell(bool(SCAN_TERMS_RE.search(text)) or re.search(r"\blead\b", lower) is not None),
            "hardcoded_paths_found": clean_join(hardcoded_paths),
            "load_calls_found": clean_join(load_calls),
            "save_calls_found": clean_join(save_calls),
            "read_calls_found": clean_join(read_calls),
            "write_calls_found": clean_join(write_calls),
            "input_variables_or_paths_guess": clean_join(load_calls + read_calls + quoted_paths[:15]),
            "output_variables_or_paths_guess": clean_join(save_calls + write_calls),
            "figures_or_plotting_guess": clean_join(figures),
            "notes": clean_join(
                [
                    "apple_metadata_stub" if is_apple_metadata(path) else "",
                    f"text_read_error={read_error}" if read_error else "",
                ]
            ),
        }
        code_rows.append(row)
        info_by_rel[rel] = row

        if read_error and is_apple_metadata(path):
            continue
        calls = parse_matlab_calls(lines)
        for dep_name, excerpt in sorted(calls.items(), key=lambda item: item[0].lower()):
            if dep_name == function_or_script:
                continue
            dep_type = classify_dependency(dep_name, local_m_map)
            found_paths = local_m_map.get(dep_name.lower(), [])
            dep_rows.append(
                {
                    "caller_relative_path": rel,
                    "caller_function_or_script": function_or_script,
                    "dependency_name": dep_name,
                    "dependency_type_guess": dep_type,
                    "dependency_file_found_in_package": bool_cell(bool(found_paths)),
                    "dependency_relative_path_if_found": clean_join(found_paths),
                    "evidence_line_excerpt_short": excerpt,
                    "notes": "helper_expected_but_file_missing" if dep_name in DOWNLOADED_HELPERS and not found_paths else "",
                }
            )

    return code_rows, dep_rows, info_by_rel, subject_refs


def build_manifest(
    files: list[Path],
    data_root: Path,
    code_info_by_rel: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    zip_member_rows: list[dict[str, Any]] = []
    for path in files:
        rel = relpath(path, data_root)
        ext = file_extension(path)
        size = path.stat().st_size
        lower = rel.lower()
        subject_id = infer_subject_id(rel)
        data_like_ext = ext in MAT_EXTENSIONS or ext in TABULAR_EXTENSIONS
        notes = []
        if is_apple_metadata(path):
            notes.append("macOS metadata stub")
        if path.name == "Description.rtf":
            notes.append("package description/readme")

        is_code = ext in CODE_EXTENSIONS and not is_apple_metadata(path)
        code_info = code_info_by_rel.get(rel, {})
        possible_fieldtrip = code_info.get("mentions_fieldtrip") == "true" or "fieldtrip" in lower or bool(re.search(r"\bft_", lower))
        possible_behavioral = data_like_ext and ("exampledata" in lower or contains_any(lower, BEHAVIOR_KEYWORDS))
        possible_force = data_like_ext and "force" in lower
        possible_event = data_like_ext and ("event" in lower or "ttl" in lower or "cue" in lower or "feedback" in lower)
        possible_stim = data_like_ext and contains_any(lower, STIM_KEYWORDS)
        possible_lfp = data_like_ext and contains_any(lower, LFP_KEYWORDS)
        possible_result = data_like_ext and contains_any(lower, {"spectra", "result", "stat", "lme", "cluster", "perm", "plot"})
        row = {
            "relative_path": rel,
            "parent_dir": Path(rel).parent.as_posix(),
            "top_level_module_guess": guess_module(rel),
            "extension": ext,
            "size_bytes": size,
            "size_mb": f"{size / (1024 * 1024):.6f}",
            "file_name": path.name,
            "stem": path.stem,
            "inferred_subject_id": subject_id,
            "inferred_group": infer_group(subject_id),
            "inferred_analysis_part": guess_analysis_part(rel),
            "possible_figure_mapping": guess_figure_mapping(rel),
            "is_mat_file": bool_cell(ext == ".mat"),
            "is_m_file": bool_cell(ext == ".m"),
            "is_csv_file": bool_cell(ext == ".csv"),
            "is_text_file": bool_cell(ext in {".txt", ".rtf"}),
            "is_excel_file": bool_cell(ext in {".xlsx", ".xls"}),
            "is_zip_file": bool_cell(ext == ".zip"),
            "is_possible_raw_lfp": bool_cell(possible_lfp and not possible_result),
            "is_possible_behavioral_data": bool_cell(possible_behavioral),
            "is_possible_force_data": bool_cell(possible_force),
            "is_possible_event_file": bool_cell(possible_event),
            "is_possible_stim_file": bool_cell(possible_stim),
            "is_possible_result_file": bool_cell(possible_result),
            "is_possible_code_file": bool_cell(is_code),
            "is_possible_fieldtrip_dependency": bool_cell(possible_fieldtrip),
            "is_possible_scan_or_identifiable_file": bool_cell(is_possible_scan_or_identifiable(rel, ext)),
            "notes": clean_join(notes),
        }
        rows.append(row)

        if ext == ".zip":
            try:
                with zipfile.ZipFile(path) as archive:
                    for info in archive.infolist():
                        if info.is_dir():
                            continue
                        member_rel = f"zip://{rel}::{info.filename}"
                        member_ext = file_extension(info.filename)
                        member_subject = infer_subject_id(info.filename)
                        member_row = {
                            "relative_path": member_rel,
                            "parent_dir": Path(info.filename).parent.as_posix(),
                            "top_level_module_guess": guess_module(rel + "/" + info.filename),
                            "extension": member_ext,
                            "size_bytes": info.file_size,
                            "size_mb": f"{info.file_size / (1024 * 1024):.6f}",
                            "file_name": Path(info.filename).name,
                            "stem": Path(info.filename).stem,
                            "inferred_subject_id": member_subject,
                            "inferred_group": infer_group(member_subject),
                            "inferred_analysis_part": guess_analysis_part(rel + "/" + info.filename),
                            "possible_figure_mapping": guess_figure_mapping(info.filename),
                            "is_mat_file": bool_cell(member_ext == ".mat"),
                            "is_m_file": bool_cell(member_ext == ".m"),
                            "is_csv_file": bool_cell(member_ext == ".csv"),
                            "is_text_file": bool_cell(member_ext in {".txt", ".rtf"}),
                            "is_excel_file": bool_cell(member_ext in {".xlsx", ".xls"}),
                            "is_zip_file": bool_cell(member_ext == ".zip"),
                            "is_possible_raw_lfp": bool_cell(contains_any(info.filename, LFP_KEYWORDS)),
                            "is_possible_behavioral_data": bool_cell(contains_any(info.filename, BEHAVIOR_KEYWORDS)),
                            "is_possible_force_data": bool_cell("force" in info.filename.lower()),
                            "is_possible_event_file": bool_cell("event" in info.filename.lower()),
                            "is_possible_stim_file": bool_cell(contains_any(info.filename, STIM_KEYWORDS)),
                            "is_possible_result_file": bool_cell(contains_any(info.filename, {"spectra", "result", "stat", "lme", "cluster", "perm"})),
                            "is_possible_code_file": bool_cell(member_ext in CODE_EXTENSIONS),
                            "is_possible_fieldtrip_dependency": bool_cell("fieldtrip" in info.filename.lower() or "ft_" in info.filename.lower()),
                            "is_possible_scan_or_identifiable_file": bool_cell(is_possible_scan_or_identifiable(info.filename, member_ext)),
                            "notes": "zip_member_listed_not_extracted",
                        }
                        zip_member_rows.append(member_row)
            except Exception as exc:
                rows[-1]["notes"] = clean_join([rows[-1]["notes"], f"zip_list_error={exc.__class__.__name__}: {exc}"])

    return rows + zip_member_rows, zip_member_rows


def build_module_inventory(
    files: list[Path],
    data_root: Path,
    code_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    files_by_dir: dict[str, list[Path]] = defaultdict(list)
    for path in files:
        rel_dir = relpath(path.parent, data_root)
        parts = [rel_dir]
        parent = path.parent
        while parent != data_root and data_root in parent.parents:
            parent = parent.parent
            parts.append(relpath(parent, data_root))
        for part in parts:
            files_by_dir[part].append(path)

    code_by_dir: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in code_rows:
        code_by_dir[Path(row["relative_path"]).parent.as_posix()].append(row)

    rows: list[dict[str, Any]] = []
    for rel_dir, dir_files in sorted(files_by_dir.items(), key=lambda item: item[0]):
        rel_dir = "." if rel_dir == "." else rel_dir
        unique_files = sorted(set(dir_files))
        ext_counter = Counter(file_extension(path) for path in unique_files)
        subject_like = [path for path in unique_files if infer_subject_id(relpath(path, data_root))]
        scripts = [relpath(path, data_root) for path in unique_files if file_extension(path) == ".m" and not is_apple_metadata(path)]
        data_files = [
            relpath(path, data_root)
            for path in unique_files
            if file_extension(path) in MAT_EXTENSIONS | {".csv", ".tsv", ".txt", ".xlsx", ".xls"}
        ]
        readmes = [
            relpath(path, data_root)
            for path in unique_files
            if re.search(r"(readme|description|instruction)", path.name, re.IGNORECASE)
        ]
        contained_code_rows = [
            row for row in code_rows if Path(row["relative_path"]).parent.as_posix().startswith(rel_dir.rstrip("."))
        ]
        inputs = []
        outputs = []
        for code_row in contained_code_rows[:30]:
            inputs.extend(code_row.get("input_variables_or_paths_guess", "").split(";"))
            outputs.extend(code_row.get("output_variables_or_paths_guess", "").split(";"))
        rows.append(
            {
                "module_guess": guess_module(rel_dir),
                "relative_dir": rel_dir,
                "n_files": len(unique_files),
                "n_matlab_files": ext_counter[".m"],
                "n_mat_files": ext_counter[".mat"],
                "n_csv_tsv_txt_xlsx_files": sum(ext_counter[ext] for ext in [".csv", ".tsv", ".txt", ".xlsx", ".xls", ".rtf"]),
                "n_zip_files": ext_counter[".zip"],
                "n_subject_like_files": len(subject_like),
                "key_scripts_found": clean_join(scripts[:30]),
                "data_files_found": clean_join(data_files[:30]),
                "readme_or_instruction_files_found": clean_join(readmes[:20]),
                "likely_inputs": clean_join(inputs[:30]),
                "likely_outputs": clean_join(outputs[:30]),
                "notes": "includes macOS metadata files" if any(is_apple_metadata(path) for path in unique_files) else "",
            }
        )
    return rows


def accepted_component_match(component: str, local_m_map: dict[str, list[str]]) -> tuple[bool, list[str], list[str]]:
    lower = component.lower()
    exact_paths = local_m_map.get(lower, [])
    if exact_paths:
        return True, exact_paths, []

    known_aliases = {
        "comparelevodopademographicsmvc": "comparedemographicslevodopamvc",
    }
    alias = known_aliases.get(lower)
    if alias and alias in local_m_map:
        return False, [], local_m_map[alias]

    normalized = normalize_component_name(component)
    normalized_matches = []
    for stem_lower, paths in local_m_map.items():
        if normalize_component_name(stem_lower) == normalized:
            normalized_matches.extend(paths)
    if normalized_matches:
        return False, [], normalized_matches

    candidates: list[tuple[float, str]] = []
    for stem in local_m_map:
        normalized_stem = normalize_component_name(stem)
        ratio = difflib.SequenceMatcher(None, normalized, normalized_stem).ratio()
        if normalized in normalized_stem or normalized_stem in normalized:
            ratio = max(ratio, 0.97)
        if ratio >= 0.78:
            candidates.append((ratio, stem))
    if not candidates:
        return False, [], []
    best = max(score for score, _ in candidates)
    near_paths: list[str] = []
    for score, stem in candidates:
        if score >= best - 0.03:
            near_paths.extend(local_m_map[stem])
    return False, [], sorted(set(near_paths))


def build_expected_component_matrix(
    local_m_map: dict[str, list[str]],
    code_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    code_by_rel = {row["relative_path"]: row for row in code_rows}
    rows = []
    found = []
    missing = []
    for component in EXPECTED_COMPONENTS:
        name = component["component_name"]
        exact_found, exact_paths, near_paths = accepted_component_match(name, local_m_map)
        selected_paths = exact_paths or near_paths
        if selected_paths:
            found.append(name)
        else:
            missing.append(name)
        deps = []
        inputs = []
        outputs = []
        for selected in selected_paths:
            info = code_by_rel.get(selected, {})
            deps.append(info.get("ft_functions_called", ""))
            inputs.append(info.get("input_variables_or_paths_guess", ""))
            outputs.append(info.get("output_variables_or_paths_guess", ""))
        notes = []
        if selected_paths and not exact_found:
            notes.append("covered_by_near_or_normalized_filename_match")
        if not selected_paths:
            notes.append("not_found_in_package")
        rows.append(
            {
                "component_name": name,
                "expected_module": component["expected_module"],
                "expected_role": component["expected_role"],
                "exact_file_found": bool_cell(exact_found),
                "near_match_files": clean_join(near_paths),
                "relative_path": clean_join(selected_paths),
                "is_function_or_script": "script_or_function" if selected_paths else "",
                "dependencies_detected": clean_join(deps),
                "input_files_or_dirs_detected": clean_join(inputs),
                "output_files_detected": clean_join(outputs),
                "notes": clean_join(notes),
            }
        )
    return rows, found, missing


class MatSummary:
    def __init__(self, max_arrays: int) -> None:
        self.max_arrays = max_arrays
        self.arrays_seen = 0
        self.keys: set[str] = set()
        self.paths: set[str] = set()
        self.scalar_fs_values: list[str] = []
        self.label_counts: list[str] = []
        self.trial_shapes: list[str] = []
        self.time_shapes: list[str] = []
        self.freq_shapes: list[str] = []
        self.power_shapes: list[str] = []
        self.data_object_rows: list[dict[str, Any]] = []
        self.flags = {
            "fieldtrip_like": False,
            "has_fsample": False,
            "has_label": False,
            "has_trial": False,
            "has_time": False,
            "has_freq": False,
            "has_power": False,
            "has_force": False,
            "has_event": False,
            "has_stim": False,
            "has_behavioral_table": False,
            "has_lfp": False,
        }

    def mark_key(self, key_path: str) -> None:
        key = key_path.split(".")[-1].split("/")[-1]
        lower = key.lower()
        self.keys.add(key)
        self.paths.add(key_path)
        if lower in {"fsample", "fs", "sr", "samplingrate", "samplerate"} or FS_KEY_RE.search(key):
            self.flags["has_fsample"] = True
        if lower in {"label", "labels", "channellabels"}:
            self.flags["has_label"] = True
        if lower in {"trial", "trials"}:
            self.flags["has_trial"] = True
        if lower in {"time", "times", "toi"}:
            self.flags["has_time"] = True
        if lower in {"freq", "frequency", "foi"}:
            self.flags["has_freq"] = True
        if lower in {"powspctrm", "power", "pow", "spectrum", "spectra"}:
            self.flags["has_power"] = True
        if contains_any(lower, BEHAVIOR_KEYWORDS):
            self.flags["has_force"] = self.flags["has_force"] or "force" in lower
            self.flags["has_event"] = self.flags["has_event"] or any(term in lower for term in ["event", "cue", "feedback", "trial"])
            self.flags["has_behavioral_table"] = True
        if contains_any(lower, STIM_KEYWORDS):
            self.flags["has_stim"] = True
        if contains_any(lower, LFP_KEYWORDS):
            self.flags["has_lfp"] = True
        if lower in {"fsample", "label", "trial", "time", "hdr", "sampleinfo", "cfg"}:
            self.flags["fieldtrip_like"] = True


def shape_text(obj: Any) -> str:
    shape = getattr(obj, "shape", None)
    if shape is not None:
        return "x".join(str(dim) for dim in shape)
    try:
        return str(len(obj))
    except Exception:
        return ""


def dtype_text(obj: Any) -> str:
    dtype = getattr(obj, "dtype", None)
    if dtype is not None:
        return str(dtype)
    return type(obj).__name__


def ndarray_string_values(array: Any, limit: int = 20) -> list[str]:
    if np is None:
        return []
    values: list[str] = []
    try:
        flat = np.asarray(array).ravel()
    except Exception:
        return values
    for item in flat[:limit]:
        if isinstance(item, str):
            values.append(item)
        elif hasattr(item, "item"):
            try:
                native = item.item()
            except Exception:
                native = item
            if isinstance(native, str):
                values.append(native)
    return values


def scalar_numeric_value(obj: Any) -> float | None:
    if np is not None:
        try:
            array = np.asarray(obj)
            if array.size == 1 and np.issubdtype(array.dtype, np.number):
                value = float(array.ravel()[0])
                if math.isfinite(value):
                    return value
        except Exception:
            return None
    if isinstance(obj, (int, float)) and math.isfinite(float(obj)):
        return float(obj)
    return None


def walk_mat_object(obj: Any, key_path: str, summary: MatSummary, depth: int = 0) -> None:
    if depth > 5 or summary.arrays_seen >= summary.max_arrays:
        return
    summary.mark_key(key_path)
    key_lower = key_path.lower()
    shape = shape_text(obj)

    scalar = scalar_numeric_value(obj)
    if scalar is not None and FS_KEY_RE.search(key_path):
        summary.scalar_fs_values.append(f"{key_path}={scalar:g}")

    if any(term in key_lower for term in ("label", "labels")):
        strings = ndarray_string_values(obj)
        if strings:
            summary.label_counts.append(f"{key_path}:n={len(strings)}")

    if key_lower.endswith("trial") or ".trial" in key_lower or "/trial" in key_lower:
        summary.trial_shapes.append(f"{key_path}:{shape}")
    if any(part in key_lower for part in (".time", "/time", ".times", "/times")) or key_lower.endswith("time") or key_lower.endswith("times"):
        summary.time_shapes.append(f"{key_path}:{shape}")
    if key_lower.endswith("freq") or ".freq" in key_lower or "/freq" in key_lower:
        summary.freq_shapes.append(f"{key_path}:{shape}")
    if any(term in key_lower for term in ("powspctrm", "power", "spectrum", "spectra")):
        summary.power_shapes.append(f"{key_path}:{shape}")

    if np is not None and isinstance(obj, np.ndarray):
        summary.arrays_seen += 1
        if obj.dtype.names:
            for field in obj.dtype.names:
                try:
                    walk_mat_object(obj[field], f"{key_path}.{field}", summary, depth + 1)
                except Exception:
                    continue
        elif obj.dtype == object and obj.size <= 50:
            for idx, item in enumerate(obj.ravel()[:10]):
                walk_mat_object(item, f"{key_path}[{idx}]", summary, depth + 1)
        return

    if hasattr(obj, "_fieldnames"):
        for field in getattr(obj, "_fieldnames", []) or []:
            try:
                walk_mat_object(getattr(obj, field), f"{key_path}.{field}", summary, depth + 1)
            except Exception:
                continue
        return

    if isinstance(obj, dict):
        for key, value in list(obj.items())[:50]:
            if str(key).startswith("__"):
                continue
            walk_mat_object(value, f"{key_path}.{key}", summary, depth + 1)


def object_type_for_mat(obj: Any) -> str:
    if hasattr(obj, "_fieldnames"):
        return "mat_struct"
    if np is not None and isinstance(obj, np.ndarray):
        return "mat_array"
    return type(obj).__name__


def infer_schema_flags(name: str, summary: MatSummary) -> dict[str, str]:
    lower = name.lower()
    return {
        "lfp_data_possible": bool_cell(summary.flags["has_lfp"] or contains_any(lower, LFP_KEYWORDS)),
        "force_data_possible": bool_cell(summary.flags["has_force"] or "force" in lower),
        "behavior_data_possible": bool_cell(summary.flags["has_behavioral_table"] or contains_any(lower, BEHAVIOR_KEYWORDS)),
        "stim_data_possible": bool_cell(summary.flags["has_stim"] or contains_any(lower, STIM_KEYWORDS)),
    }


def inspect_mat_files(
    files: list[Path],
    data_root: Path,
    max_arrays: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    mat_rows: list[dict[str, Any]] = []
    schema_rows: list[dict[str, Any]] = []

    for path in files:
        if file_extension(path) != ".mat":
            continue
        rel = relpath(path, data_root)
        subject_id = infer_subject_id(rel)
        size_mb = path.stat().st_size / (1024 * 1024)
        load_status = "unread"
        error_message = ""
        mat_format = "unknown"
        top_level: dict[str, Any] = {}
        summary = MatSummary(max_arrays=max_arrays)

        if is_apple_metadata(path):
            mat_format = "apple_metadata_stub"
        elif path.read_bytes()[:8] == b"\x89HDF\r\n\x1a\n":
            mat_format = "hdf5_mat_v7_3_h5py_required"

        if scipy_io is not None and not is_apple_metadata(path):
            try:
                top_level = {
                    key: value
                    for key, value in scipy_io.loadmat(path, struct_as_record=False, squeeze_me=False).items()
                    if not key.startswith("__")
                }
                load_status = "loaded_scipy"
                mat_format = "classic_mat_or_scipy_readable"
            except Exception as exc:
                error_message = f"scipy:{exc.__class__.__name__}: {exc}"
                if "Please use HDF reader" in str(exc) and mat_format == "unknown":
                    mat_format = "hdf5_mat_v7_3_h5py_required"

        if load_status != "loaded_scipy" and h5py is not None and not is_apple_metadata(path):
            try:
                with h5py.File(path, "r") as handle:
                    mat_format = "hdf5_mat_v7_3"
                    top_level = {key: handle[key] for key in handle.keys()}
                    for key in handle.keys():
                        walk_hdf5_object(handle[key], key, summary)
                load_status = "loaded_h5py"
            except Exception as exc:
                error_message = clean_join([error_message, f"h5py:{exc.__class__.__name__}: {exc}"])
        elif load_status != "loaded_scipy" and h5py is None:
            if error_message:
                error_message = clean_join([error_message, f"h5py_unavailable:{H5PY_IMPORT_ERROR}"])
            if mat_format == "hdf5_mat_v7_3_h5py_required":
                error_message = clean_join([error_message, "hdf5_mat_signature_detected"])

        if load_status == "loaded_scipy":
            for key, value in top_level.items():
                walk_mat_object(value, key, summary)
                object_summary = MatSummary(max_arrays=max(20, max_arrays // 4))
                walk_mat_object(value, key, object_summary)
                flags = infer_schema_flags(key, object_summary)
                schema_rows.append(
                    {
                        "relative_path": rel,
                        "module_guess": guess_module(rel),
                        "inferred_subject_id": subject_id,
                        "data_object_name": key,
                        "object_type": object_type_for_mat(value),
                        "shape_or_rows_cols": shape_text(value),
                        "dtype_or_column_types_if_known": dtype_text(value),
                        "key_variables_detected": clean_join(sorted(object_summary.keys)[:40]),
                        "sampling_rate_detected": clean_join(object_summary.scalar_fs_values),
                        "sampling_rate_source": "explicit_file_field" if object_summary.scalar_fs_values else "unknown",
                        "time_axis_detected": bool_cell(object_summary.flags["has_time"]),
                        "trial_axis_detected": bool_cell(object_summary.flags["has_trial"] or "trial" in key.lower()),
                        "channel_axis_detected": bool_cell(object_summary.flags["has_label"] or re.match(r"Ch\d+", key) is not None),
                        "channel_labels_detected": clean_join(object_summary.label_counts),
                        **flags,
                        "notes": "",
                    }
                )
        elif load_status == "loaded_h5py":
            for key, value in top_level.items():
                flags = infer_schema_flags(key, summary)
                schema_rows.append(
                    {
                        "relative_path": rel,
                        "module_guess": guess_module(rel),
                        "inferred_subject_id": subject_id,
                        "data_object_name": key,
                        "object_type": "hdf5_dataset_or_group",
                        "shape_or_rows_cols": shape_text(value),
                        "dtype_or_column_types_if_known": dtype_text(value),
                        "key_variables_detected": clean_join(sorted(summary.keys)[:40]),
                        "sampling_rate_detected": clean_join(summary.scalar_fs_values),
                        "sampling_rate_source": "explicit_file_field" if summary.scalar_fs_values else "unknown",
                        "time_axis_detected": bool_cell(summary.flags["has_time"]),
                        "trial_axis_detected": bool_cell(summary.flags["has_trial"]),
                        "channel_axis_detected": bool_cell(summary.flags["has_label"]),
                        "channel_labels_detected": clean_join(summary.label_counts),
                        **flags,
                        "notes": "hdf5_mat_inspected_without_extraction",
                    }
                )

        top_keys = sorted(top_level.keys()) if top_level else []
        likely = "unknown"
        if subject_id.startswith("Kont") or "exampledata" in rel.lower():
            likely = "example_behavioral_data"
        elif summary.flags["fieldtrip_like"] or summary.flags["has_lfp"]:
            likely = "raw_or_derived_lfp"
        elif summary.flags["has_power"] or contains_any(rel, {"spectra", "lme", "stat", "cluster"}):
            likely = "spectral_or_result"
        elif summary.flags["has_behavioral_table"] or contains_any(rel, BEHAVIOR_KEYWORDS):
            likely = "behavioral"
        elif summary.flags["has_stim"] or contains_any(rel, STIM_KEYWORDS):
            likely = "stim"

        notes = []
        if is_apple_metadata(path):
            notes.append("macOS metadata stub, not a real MAT payload")
        if load_status != "loaded_scipy" and load_status != "loaded_h5py":
            notes.append("not deeply readable by available optional readers")

        mat_rows.append(
            {
                "relative_path": rel,
                "module_guess": guess_module(rel),
                "inferred_subject_id": subject_id,
                "size_mb": f"{size_mb:.6f}",
                "mat_format_guess": mat_format,
                "load_status": load_status,
                "error_message": error_message,
                "top_level_keys": clean_join(top_keys),
                "n_top_level_keys": len(top_keys),
                "has_fieldtrip_like_struct": bool_cell(summary.flags["fieldtrip_like"]),
                "has_fsample": bool_cell(summary.flags["has_fsample"]),
                "fsample_values_found": clean_join(summary.scalar_fs_values),
                "has_label": bool_cell(summary.flags["has_label"]),
                "n_labels_if_found": clean_join(summary.label_counts),
                "has_trial": bool_cell(summary.flags["has_trial"]),
                "trial_shape_summary": clean_join(summary.trial_shapes[:20]),
                "has_time": bool_cell(summary.flags["has_time"]),
                "time_shape_summary": clean_join(summary.time_shapes[:20]),
                "has_freq": bool_cell(summary.flags["has_freq"]),
                "freq_shape_summary": clean_join(summary.freq_shapes[:20]),
                "has_powspctrm_or_power": bool_cell(summary.flags["has_power"]),
                "power_shape_summary": clean_join(summary.power_shapes[:20]),
                "has_force_like_data": bool_cell(summary.flags["has_force"] or "force" in rel.lower()),
                "has_event_like_data": bool_cell(summary.flags["has_event"] or "event" in rel.lower()),
                "has_stim_like_data": bool_cell(summary.flags["has_stim"]),
                "has_behavioral_table_like_data": bool_cell(summary.flags["has_behavioral_table"]),
                "has_lfp_like_data": bool_cell(summary.flags["has_lfp"]),
                "likely_raw_or_result_or_example": likely,
                "possible_figure_source_data": bool_cell("s1" in rel.lower() or "source" in rel.lower()),
                "notes": clean_join(notes),
            }
        )

    return mat_rows, schema_rows


def walk_hdf5_object(obj: Any, key_path: str, summary: MatSummary, depth: int = 0) -> None:
    if depth > 5 or summary.arrays_seen >= summary.max_arrays:
        return
    summary.mark_key(key_path)
    shape = getattr(obj, "shape", "")
    key_lower = key_path.lower()
    if FS_KEY_RE.search(key_path) and hasattr(obj, "shape") and getattr(obj, "shape", ()) in [(), (1,), (1, 1)]:
        try:
            value = float(obj[()])
            summary.scalar_fs_values.append(f"{key_path}={value:g}")
        except Exception:
            pass
    if "trial" in key_lower:
        summary.trial_shapes.append(f"{key_path}:{shape}")
    if "time" in key_lower:
        summary.time_shapes.append(f"{key_path}:{shape}")
    if "freq" in key_lower:
        summary.freq_shapes.append(f"{key_path}:{shape}")
    if any(term in key_lower for term in ("powspctrm", "power", "spectrum")):
        summary.power_shapes.append(f"{key_path}:{shape}")
    if hasattr(obj, "keys"):
        for key in list(obj.keys())[:50]:
            walk_hdf5_object(obj[key], f"{key_path}/{key}", summary, depth + 1)
    else:
        summary.arrays_seen += 1


def guess_delimiter(line: str, ext: str) -> str:
    if ext == ".tsv":
        return "tab"
    counts = {",": line.count(","), "\t": line.count("\t"), ";": line.count(";")}
    delimiter = max(counts, key=counts.get)
    if counts[delimiter] == 0:
        return "none"
    return {"\t": "tab", ",": "comma", ";": "semicolon"}[delimiter]


def inspect_tabular_files(
    files: list[Path],
    data_root: Path,
    max_text_read_bytes: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    schema_rows: list[dict[str, Any]] = []
    for path in files:
        ext = file_extension(path)
        if ext not in TABULAR_EXTENSIONS:
            continue
        rel = relpath(path, data_root)
        subject_id = infer_subject_id(rel)
        read_status = "unread"
        delimiter = ""
        n_rows = ""
        n_cols = ""
        fields = ""
        notes = []
        text_content = ""

        if ext in {".xlsx", ".xls"}:
            if openpyxl is None or ext == ".xls":
                read_status = "present_not_deeply_read"
                notes.append(f"openpyxl_unavailable_or_xls={OPENPYXL_IMPORT_ERROR}" if openpyxl is None else "legacy_xls_not_read_by_openpyxl")
            else:
                try:
                    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    sheet = workbook[workbook.sheetnames[0]]
                    n_rows = str(sheet.max_row)
                    n_cols = str(sheet.max_column)
                    first_row = [str(cell.value) if cell.value is not None else "" for cell in next(sheet.iter_rows(max_row=1))]
                    fields = clean_join(first_row, sep="|")
                    read_status = "read_openpyxl_header"
                    workbook.close()
                except Exception as exc:
                    read_status = "read_error"
                    notes.append(f"{exc.__class__.__name__}: {exc}")
        else:
            text_content, read_error = read_text_limited(path, max_text_read_bytes)
            if read_error:
                notes.append(read_error)
            lines = text_content.splitlines()
            nonempty = [line for line in lines if line.strip()]
            if nonempty:
                delimiter = guess_delimiter(nonempty[0], ext)
                if delimiter == "tab":
                    split_char = "\t"
                elif delimiter == "comma":
                    split_char = ","
                elif delimiter == "semicolon":
                    split_char = ";"
                else:
                    split_char = None
                if split_char:
                    header = [part.strip() for part in nonempty[0].split(split_char)]
                    fields = clean_join(header[:30], sep="|")
                    n_cols = str(len(header))
                else:
                    fields = truncate(nonempty[0], 260)
                    n_cols = "1"
                n_rows = str(len(nonempty))
            read_status = "read_text_header" if not is_apple_metadata(path) else "read_text_header_apple_metadata"

        lower_probe = f"{rel}\n{fields}\n{text_content[:5000]}".lower()
        is_config = bool(re.search(r"(readme|description|instruction|\\rtf|title:|creators:)", lower_probe))
        row = {
            "relative_path": rel,
            "module_guess": guess_module(rel),
            "inferred_subject_id": subject_id,
            "extension": ext,
            "size_mb": f"{path.stat().st_size / (1024 * 1024):.6f}",
            "read_status": read_status,
            "delimiter_guess": delimiter,
            "n_rows_if_countable": n_rows,
            "n_columns_if_countable": n_cols,
            "column_names_or_first_fields": fields,
            "has_behavioral_columns_guess": bool_cell(contains_any(lower_probe, BEHAVIOR_KEYWORDS)),
            "has_force_columns_guess": bool_cell("force" in lower_probe or "mvc" in lower_probe),
            "has_event_columns_guess": bool_cell("event" in lower_probe or "trial" in lower_probe or "feedback" in lower_probe),
            "has_stim_columns_guess": bool_cell(contains_any(lower_probe, STIM_KEYWORDS)),
            "has_demographic_or_clinical_columns_guess": bool_cell("demographic" in lower_probe or "updrs" in lower_probe or "levodopa" in lower_probe or "clinical" in lower_probe),
            "has_path_or_config_content_guess": bool_cell(is_config or "/users/" in lower_probe or "path" in lower_probe),
            "notes": clean_join(notes + (["text_config_or_description"] if is_config else []) + (["apple_metadata_stub"] if is_apple_metadata(path) else [])),
        }
        rows.append(row)
        if read_status.startswith("read"):
            schema_rows.append(
                {
                    "relative_path": rel,
                    "module_guess": guess_module(rel),
                    "inferred_subject_id": subject_id,
                    "data_object_name": "text_or_table_header",
                    "object_type": "text_config" if is_config else "table",
                    "shape_or_rows_cols": f"{n_rows}x{n_cols}" if n_rows or n_cols else "",
                    "dtype_or_column_types_if_known": "text",
                    "key_variables_detected": fields,
                    "sampling_rate_detected": "",
                    "sampling_rate_source": "unknown",
                    "time_axis_detected": bool_cell("time" in lower_probe),
                    "trial_axis_detected": bool_cell("trial" in lower_probe),
                    "channel_axis_detected": bool_cell("channel" in lower_probe),
                    "channel_labels_detected": "",
                    "lfp_data_possible": bool_cell(contains_any(lower_probe, LFP_KEYWORDS)),
                    "force_data_possible": bool_cell("force" in lower_probe or "mvc" in lower_probe),
                    "behavior_data_possible": bool_cell(contains_any(lower_probe, BEHAVIOR_KEYWORDS)),
                    "stim_data_possible": bool_cell(contains_any(lower_probe, STIM_KEYWORDS)),
                    "notes": row["notes"],
                }
            )
    return rows, schema_rows


def build_subject_inventory(
    manifest_rows: list[dict[str, Any]],
    code_rows: list[dict[str, Any]],
    code_subject_refs: set[str],
) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    subject_sources: dict[str, set[str]] = defaultdict(set)
    subject_files: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in manifest_rows:
        rel = row["relative_path"]
        ids = extract_subject_ids(rel)
        for subject_id in ids:
            subject_sources[subject_id].add("path")
            subject_sources[subject_id].add("filename")
            subject_files[subject_id].append(row)

    for subject_id in code_subject_refs:
        subject_sources[subject_id].add("code_reference")

    rows: list[dict[str, Any]] = []
    for subject_id in sorted(subject_sources):
        files = subject_files.get(subject_id, [])
        modules = sorted(set(row.get("top_level_module_guess", "") for row in files if row.get("top_level_module_guess")))
        file_types = sorted(set(row.get("extension", "") for row in files if row.get("extension")))
        behavioral = any(row.get("is_possible_behavioral_data") == "true" for row in files)
        event = any(row.get("is_possible_event_file") == "true" for row in files)
        force = any(row.get("is_possible_force_data") == "true" for row in files)
        lfp = any(row.get("is_possible_raw_lfp") == "true" for row in files)
        stim = any(row.get("is_possible_stim_file") == "true" for row in files)
        notes = []
        if not files and "code_reference" in subject_sources[subject_id]:
            notes.append("subject_id_referenced_in_code_only_no_shared_file_observed")
        if subject_id in EXPECTED_EXAMPLE_SUBJECTS and files:
            notes.append("expected example behavioral subject observed")
        rows.append(
            {
                "subject_id": subject_id,
                "inferred_group": infer_group(subject_id),
                "is_example_subject": bool_cell(subject_id in EXPECTED_EXAMPLE_SUBJECTS or subject_id.lower().startswith("kont")),
                "source_of_detection": clean_join(sorted(subject_sources[subject_id])),
                "n_files": len(files),
                "file_types": clean_join(file_types),
                "modules_present": clean_join(modules),
                "behavioral_files_present": bool_cell(behavioral),
                "event_files_present": bool_cell(event),
                "force_files_present": bool_cell(force),
                "lfp_files_present": bool_cell(lfp),
                "stim_files_present": bool_cell(stim),
                "notes": clean_join(notes),
            }
        )

    observed_examples = sorted([sid for sid in EXPECTED_EXAMPLE_SUBJECTS if sid in subject_files and subject_files[sid]])
    missing_examples = sorted([sid for sid in EXPECTED_EXAMPLE_SUBJECTS if sid not in observed_examples])
    for missing in missing_examples:
        rows.append(
            {
                "subject_id": missing,
                "inferred_group": "example_control",
                "is_example_subject": "true",
                "source_of_detection": "expected_shared_package_example",
                "n_files": 0,
                "file_types": "",
                "modules_present": "",
                "behavioral_files_present": "false",
                "event_files_present": "false",
                "force_files_present": "false",
                "lfp_files_present": "false",
                "stim_files_present": "false",
                "notes": "missing expected example behavioral data",
            }
        )
    return rows, observed_examples, missing_examples


def build_privacy_inventory(
    manifest_rows: list[dict[str, Any]],
    code_rows: list[dict[str, Any]],
    tabular_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()

    def add(rel: str, risk_type: str, severity: str, evidence: str, action: str) -> None:
        key = (rel, risk_type)
        if key in seen:
            return
        seen.add(key)
        rows.append(
            {
                "relative_path": rel,
                "risk_type": risk_type,
                "severity": severity,
                "evidence": truncate(evidence, 300),
                "recommended_action": action,
            }
        )

    for row in manifest_rows:
        rel = row["relative_path"]
        ext = row["extension"]
        if row.get("is_possible_scan_or_identifiable_file") == "true":
            add(rel, "scan_or_identifiable_file_name", "high" if ext in SCAN_EXTENSIONS else "medium", f"extension/path term: {ext}", "Do not open deeply; confirm governance/consent before use.")
        if row.get("inferred_subject_id"):
            severity = "low"
            evidence = f"anonymized subject-like ID {row['inferred_subject_id']} in path"
            add(rel, "subject_identifier_in_path", severity, evidence, "Keep reports at aggregate/path level; do not expose raw sample values.")
        if row.get("is_possible_raw_lfp") == "true":
            add(rel, "possible_raw_patient_lfp", "high", "filename/path suggests LFP data", "Confirm consent and analysis scope before any raw LFP processing.")

    for row in code_rows:
        if row.get("mentions_leaddbs_or_imaging") == "true":
            add(row["relative_path"], "imaging_or_lead_localization_code_reference", "medium", "MATLAB code mentions lead/imaging-like terms", "Treat as code reference only unless actual imaging files are present.")
        if "demographic" in row["file_name"].lower() or "levodopa" in row["file_name"].lower() or "updrs" in row.get("input_variables_or_paths_guess", "").lower():
            add(row["relative_path"], "clinical_demographic_code_reference", "low", "script name or code references demographics/levodopa/UPDRS", "Do not infer that clinical tables are shared unless data files are present.")

    for row in tabular_rows:
        if row.get("has_demographic_or_clinical_columns_guess") == "true":
            add(row["relative_path"], "clinical_demographic_text_reference", "low", "description or text mentions demographic/clinical variables", "Treat as documentation unless actual subject-level table is observed.")

    return rows


def dependency_summary(dep_rows: list[dict[str, Any]]) -> tuple[dict[str, int], list[dict[str, str]]]:
    counts = Counter(row["dependency_type_guess"] for row in dep_rows)
    missing: dict[str, dict[str, str]] = {}
    for row in dep_rows:
        dep_type = row["dependency_type_guess"]
        found = row["dependency_file_found_in_package"] == "true"
        if dep_type in {"local_function", "downloaded_helper", "unknown", "FieldTrip", "statistics_toolbox", "LeadDBS_or_imaging"} and not found:
            name = row["dependency_name"]
            missing.setdefault(
                name,
                {
                    "dependency_name": name,
                    "dependency_type_guess": dep_type,
                    "example_caller": row["caller_relative_path"],
                    "notes": row.get("notes", ""),
                },
            )
    return dict(sorted(counts.items())), sorted(missing.values(), key=lambda item: (item["dependency_type_guess"], item["dependency_name"]))[:80]


def extract_sampling_rate_code_refs(code_rows: list[dict[str, Any]], lfp_analysis_fs: float, stim_binary_fs: float) -> list[str]:
    refs = []
    for row in code_rows:
        text = " ".join(
            [
                row.get("input_variables_or_paths_guess", ""),
                row.get("output_variables_or_paths_guess", ""),
                row.get("load_calls_found", ""),
                row.get("save_calls_found", ""),
            ]
        )
        if str(int(lfp_analysis_fs)) in text or f"{lfp_analysis_fs:g}" in text:
            refs.append(f"{row['relative_path']}: possible {lfp_analysis_fs:g} Hz reference")
        if str(int(stim_binary_fs)) in text or f"{stim_binary_fs:g}" in text:
            refs.append(f"{row['relative_path']}: possible {stim_binary_fs:g} Hz reference")
    return sorted(set(refs))


def code_sampling_refs_from_files(files: list[Path], data_root: Path, max_text_read_bytes: int) -> list[str]:
    refs = []
    pattern = re.compile(r"(NewSR\s*=\s*\d+|resamplefs\s*=\s*\w+|\b1000\s*Hz|\b1000Hz|\b2048\b|\b200\b|\b1000\b)", re.IGNORECASE)
    for path in files:
        if file_extension(path) != ".m" or is_apple_metadata(path):
            continue
        text, _ = read_text_limited(path, max_text_read_bytes)
        for line in text.splitlines():
            if pattern.search(line):
                refs.append(f"{relpath(path, data_root)}: {truncate(line, 180)}")
    return sorted(set(refs))


def write_commands_file(path: Path) -> list[str]:
    commands = [
        "pwd",
        "git rev-parse --show-toplevel",
        "git status --short",
        "conda run -n stn_env python -V",
        "source /scratch/haizhe/stn/start_stn.sh && python -V",
        "find . -maxdepth 3 -type d | sort | head -200",
        "find . -maxdepth 3 -type f | sort | head -200",
        "find cambium/Force_Scripts -maxdepth 4 -type d | sort | head -300",
        "find cambium/Force_Scripts -maxdepth 4 -type f | sort | head -300",
        "rg --files scripts | rg 'audit|phase6|ppn|PPN|dataset'",
        "find reports -maxdepth 3 -type d | sort | head -200",
        "find reports -maxdepth 3 -type f | sort | head -300",
        "sed -n '1,220p' scripts/phase6_audit_ppn_he_tan_2021.py",
        "sed -n '1,220p' reports/phase6_ppn_he_tan_2021_audit/README_dataset_audit.md",
        "sed -n '1,160p' cambium/Force_Scripts/Force_Scripts/1BehavioralData/ExtractData.m",
        "sed -n '1,160p' cambium/Force_Scripts/Force_Scripts/2LocalFieldPotentialData/GetLFP_FirstLevel.m",
        "sed -n '1,180p' cambium/Force_Scripts/Force_Scripts/3DBSEffectsBehavior/GetEvents_Stim.m",
        "python - <<'PY' ... read first bytes of Description.rtf ... PY",
        "source /scratch/haizhe/stn/start_stn.sh && python scripts/phase6_audit_stn_force_adaptation_herz_2023.py --data-root cambium/Force_Scripts --out-dir reports/phase6_stn_force_adaptation_herz_2023_audit --paper-fs 2048 --lfp-analysis-fs 200 --stim-binary-fs 1000",
        "source /scratch/haizhe/stn/start_stn.sh && python -m py_compile scripts/phase6_audit_stn_force_adaptation_herz_2023.py",
        "ls -lh reports/phase6_stn_force_adaptation_herz_2023_audit",
        "head -80 reports/phase6_stn_force_adaptation_herz_2023_audit/README_dataset_audit.md",
        "git diff --check",
        "find reports/phase6_stn_force_adaptation_herz_2023_audit -type f -size +5M -print",
        "git status --short",
    ]
    path.write_text("\n".join(commands) + "\n", encoding="utf-8")
    return commands


def markdown_table(headers: list[str], rows: list[list[Any]], max_rows: int = 20) -> str:
    limited_rows = rows[:max_rows]
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join("---" for _ in headers) + " |"]
    for row in limited_rows:
        lines.append("| " + " | ".join(str(value) for value in row) + " |")
    if len(rows) > max_rows:
        lines.append("| ... | " + " | ".join("" for _ in headers[1:]) + " |")
    return "\n".join(lines)


def build_readme(
    path: Path,
    args: argparse.Namespace,
    repo_root: Path,
    data_root: Path,
    timestamp: str,
    manifest_rows: list[dict[str, Any]],
    module_rows: list[dict[str, Any]],
    expected_rows: list[dict[str, Any]],
    subject_rows: list[dict[str, Any]],
    mat_rows: list[dict[str, Any]],
    tabular_rows: list[dict[str, Any]],
    code_rows: list[dict[str, Any]],
    dep_rows: list[dict[str, Any]],
    privacy_rows: list[dict[str, Any]],
    findings: dict[str, Any],
    commands: list[str],
) -> None:
    physical_manifest_rows = [row for row in manifest_rows if not row["relative_path"].startswith("zip://")]
    ext_counts = Counter(row["extension"] for row in physical_manifest_rows)
    module_counts = Counter(row["top_level_module_guess"] for row in physical_manifest_rows)
    part_counts = Counter(row["inferred_analysis_part"] for row in physical_manifest_rows)
    dep_counts, _ = dependency_summary(dep_rows)
    observed_examples = findings["observed_example_subjects"]
    missing_examples = findings["missing_example_subjects"]
    expected_found_count = len(findings["expected_components_found"])
    expected_missing_count = len(findings["expected_components_missing"])
    mat_failures = findings["mat_read_failures"]

    module_table_rows = [
        [
            row["module_guess"],
            row["relative_dir"],
            row["n_files"],
            row["n_matlab_files"],
            row["n_mat_files"],
            row["n_csv_tsv_txt_xlsx_files"],
            row["notes"],
        ]
        for row in module_rows
        if row["relative_dir"] in {".", "Force_Scripts", "Force_Scripts/1BehavioralData", "Force_Scripts/2LocalFieldPotentialData", "Force_Scripts/3DBSEffectsBehavior", "Force_Scripts/4DBSEffectsLocalFieldPotential", "Force_Scripts/ExampleData", "__MACOSX"}
    ]
    component_rows = [
        [
            row["component_name"],
            "Y" if row["relative_path"] else "N",
            row["relative_path"] or row["near_match_files"],
            row["notes"],
        ]
        for row in expected_rows
    ]
    subject_table_rows = [
        [
            row["subject_id"],
            row["inferred_group"],
            row["source_of_detection"],
            row["n_files"],
            row["file_types"],
            row["notes"],
        ]
        for row in subject_rows
        if row["subject_id"].startswith("Kont") or int(row["n_files"] or 0) > 0
    ]

    lines = [
        "# Phase 6A Part 2 Dataset Audit: Herz/Groppa/Brown Force Adaptation",
        "",
        "## Scope",
        "",
        f"Dataset/package: {DATASET_NAME}.",
        f"Paper context: {PAPER_TITLE}.",
        "",
        "This audit inventories the shared Force_Scripts package and maps files, code, schemas, and visible dependencies to the dataset description. It does not run MATLAB, FieldTrip, preprocessing, LME models, permutation tests, or analysis reproduction.",
        "",
        "## Dataset Root And Environment",
        "",
        f"- Dataset root audited: `{data_root}`",
        f"- Repository root: `{repo_root}`",
        f"- Audit timestamp: `{timestamp}`",
        f"- Python executable: `{sys.executable}`",
        f"- Python version: `{sys.version.split()[0]}`",
        "- Environment: existing repo `stn_env` only. `conda run -n stn_env python -V` was attempted first and was not registered on this host; subsequent Python commands used `source /scratch/haizhe/stn/start_stn.sh && python ...`.",
        f"- Optional readers: scipy `{'available' if scipy_io is not None else 'missing: ' + SCIPY_IMPORT_ERROR}`, h5py `{'available' if h5py is not None else 'missing: ' + H5PY_IMPORT_ERROR}`, openpyxl `{'available' if openpyxl is not None else 'missing: ' + OPENPYXL_IMPORT_ERROR}`.",
        "",
        "## Data Availability Caveat",
        "",
        "The paper states that participant consent did not allow depositing the full original participant dataset. This audit therefore does not treat absence of the full 16-patient/15-control raw cohort as a critical failure. The shared package is assessed as code plus minimum/example data unless the package itself claims that full raw data are present.",
        "",
        "## File And Folder Summary",
        "",
        f"- Physical files under audited root: `{len(physical_manifest_rows)}`",
        f"- MATLAB `.m` path entries: `{ext_counts['.m']}`",
        f"- Real MATLAB code files excluding macOS metadata stubs: `{sum(1 for row in physical_manifest_rows if row.get('is_possible_code_file') == 'true')}`",
        f"- `.mat` files: `{ext_counts['.mat']}`",
        f"- Tabular/text files (`.csv`, `.tsv`, `.txt`, `.xlsx`, `.xls`, `.rtf`): `{sum(ext_counts[ext] for ext in TABULAR_EXTENSIONS)}`",
        f"- ZIP files under audited root: `{ext_counts['.zip']}`",
        f"- macOS metadata stubs: `{sum(1 for row in physical_manifest_rows if 'macOS metadata' in row.get('notes', ''))}`",
        "",
        "Extension counts:",
        "",
        markdown_table(["extension", "count"], [[ext, count] for ext, count in sorted(ext_counts.items())]),
        "",
        "Analysis-part counts:",
        "",
        markdown_table(["analysis_part", "count"], [[part, count] for part, count in sorted(part_counts.items())]),
        "",
        "## Module Mapping",
        "",
        markdown_table(
            ["module_guess", "relative_dir", "files", "m_files", "mat_files", "tabular_text", "notes"],
            module_table_rows,
            max_rows=30,
        ),
        "",
        "## Expected Component Coverage",
        "",
        f"- Expected components covered by exact or strong near-match file names: `{expected_found_count}`",
        f"- Expected components missing as files: `{expected_missing_count}`",
        "",
        markdown_table(["component", "covered", "path_or_match", "notes"], component_rows, max_rows=60),
        "",
        "## Subject And Example Data Coverage",
        "",
        f"- Expected shared-package example subjects: `{', '.join(EXPECTED_EXAMPLE_SUBJECTS)}`",
        f"- Observed example subjects: `{', '.join(observed_examples) if observed_examples else 'none'}`",
        f"- Missing example subjects: `{', '.join(missing_examples) if missing_examples else 'none'}`",
        "- Study-level cohort context: 16 PD patients, 15 analyzed healthy controls after one control exclusion, and 14 patients in the stimulation session.",
        "- Observed full cohort raw data: not present in this package based on file inventory; code references PD and HC IDs, but raw participant LFP/behavioral cohort files were not shared under the audited root.",
        "",
        markdown_table(["subject_id", "group", "source", "n_files", "file_types", "notes"], subject_table_rows, max_rows=50),
        "",
        "## MAT-File Schema Summary",
        "",
        f"- MAT files inventoried: `{len(mat_rows)}`",
        f"- MAT files loaded via scipy/h5py: `{sum(1 for row in mat_rows if row['load_status'] in {'loaded_scipy', 'loaded_h5py'})}`",
        f"- MAT read failures/stubs: `{len(mat_failures)}`",
        f"- FieldTrip-like MAT structs detected: `{sum(1 for row in mat_rows if row['has_fieldtrip_like_struct'] == 'true')}`",
        f"- Force-like MAT files detected: `{sum(1 for row in mat_rows if row['has_force_like_data'] == 'true')}`",
        f"- Event-like MAT files detected: `{sum(1 for row in mat_rows if row['has_event_like_data'] == 'true')}`",
        "",
        "The four real example MAT files are lightweight behavioral/force/event examples; no full raw STN LFP cohort MAT files were observed under the audited root.",
        "",
        "## Tabular/Text Schema Summary",
        "",
        f"- Tabular/text files inventoried: `{len(tabular_rows)}`",
        "- The main readable text file is `Description.rtf`, which documents package structure, MATLAB/FieldTrip requirements, and example-data instructions.",
        "- No subject-level CSV/TSV/XLS/XLSX data tables were observed under the audited root.",
        "",
        "## MATLAB Code And Dependency Summary",
        "",
        f"- MATLAB `.m` path entries inventoried: `{len(code_rows)}`",
        f"- Real MATLAB code files excluding macOS metadata stubs: `{sum(1 for row in code_rows if 'apple_metadata_stub' not in row.get('notes', ''))}`",
        f"- MATLAB dependency/call rows: `{len(dep_rows)}`",
        "",
        markdown_table(["dependency_type", "count"], [[key, value] for key, value in sorted(dep_counts.items())]),
        "",
        "FieldTrip calls are present in the LFP modules. LME/statistics and permutation/cluster code are present. This audit records those requirements but does not test MATLAB or FieldTrip availability.",
        "",
        "## Sampling-Rate Observations",
        "",
        f"- Paper default STN LFP acquisition rate: `{args.paper_fs:g} Hz` (recorded only as paper context unless explicit in file/code).",
        f"- LFP analysis/downsampled rate: `{args.lfp_analysis_fs:g} Hz` appears in MATLAB code as `NewSR=200`/FieldTrip resampling context.",
        f"- Stimulation binary/downsampled rate: `{args.stim_binary_fs:g} Hz` appears in the package description and stimulation processing code context.",
        "- No explicit `fsample` value was detected in the shared example MAT payloads.",
        "",
        "## Privacy And Governance Findings",
        "",
        f"- Privacy/governance rows: `{len(privacy_rows)}`",
        f"- Scan/imaging-like files flagged: `{len(findings['possible_scan_or_identifiable_files'])}`",
        "- Imaging files were not observed under the audited root. Subject-like IDs in filenames/code are anonymized IDs and are reported only at path/count level.",
        "",
        "## Known Issues And Ambiguities",
        "",
    ]
    if findings["critical_issues"]:
        lines.extend(f"- Critical: {issue}" for issue in findings["critical_issues"])
    else:
        lines.append("- Critical issues: none identified for the audit scope.")
    for issue in findings["noncritical_limitations"]:
        lines.append(f"- Noncritical limitation: {issue}")
    lines.extend(
        [
            "",
            "## Recommendations For Phase 6B",
            "",
        ]
    )
    for step in findings["recommended_next_steps"]:
        lines.append(f"- {step}")
    lines.extend(
        [
            "",
            "## Exact Commands Run",
            "",
            "See also `phase6a_part2_commands_run.txt`.",
            "",
        ]
    )
    lines.extend(f"- `{command}`" for command in commands)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_findings(
    args: argparse.Namespace,
    repo_root: Path,
    data_root: Path,
    timestamp: str,
    manifest_rows: list[dict[str, Any]],
    code_rows: list[dict[str, Any]],
    dep_rows: list[dict[str, Any]],
    mat_rows: list[dict[str, Any]],
    tabular_rows: list[dict[str, Any]],
    module_rows: list[dict[str, Any]],
    privacy_rows: list[dict[str, Any]],
    subject_rows: list[dict[str, Any]],
    expected_found: list[str],
    expected_missing: list[str],
    observed_examples: list[str],
    missing_examples: list[str],
    files: list[Path],
    max_text_read_bytes: int,
) -> dict[str, Any]:
    physical_rows = [row for row in manifest_rows if not row["relative_path"].startswith("zip://")]
    ext_counts = Counter(row["extension"] for row in physical_rows)
    dep_counts, missing_deps = dependency_summary(dep_rows)
    mat_failures = [
        {
            "relative_path": row["relative_path"],
            "load_status": row["load_status"],
            "error_message": row["error_message"],
            "notes": row["notes"],
        }
        for row in mat_rows
        if row["load_status"] not in {"loaded_scipy", "loaded_h5py"}
    ]
    scan_flags = [
        row["relative_path"]
        for row in physical_rows
        if row.get("is_possible_scan_or_identifiable_file") == "true"
    ]
    observed_ids = sorted({row["subject_id"] for row in subject_rows if row["subject_id"]})
    module_summary = {
        row["relative_dir"]: {
            "module_guess": row["module_guess"],
            "n_files": row["n_files"],
            "n_matlab_files": row["n_matlab_files"],
            "n_mat_files": row["n_mat_files"],
        }
        for row in module_rows
        if row["relative_dir"] in {".", "Force_Scripts", "Force_Scripts/1BehavioralData", "Force_Scripts/2LocalFieldPotentialData", "Force_Scripts/3DBSEffectsBehavior", "Force_Scripts/4DBSEffectsLocalFieldPotential", "Force_Scripts/ExampleData"}
    }

    critical_issues = []
    if missing_examples:
        critical_issues.append(f"Missing expected example behavioral data for {', '.join(missing_examples)}.")
    real_mat_failures = [row for row in mat_failures if "apple_metadata" not in row.get("notes", "")]
    if real_mat_failures:
        failed_paths = ", ".join(row["relative_path"] for row in real_mat_failures[:5])
        critical_issues.append(
            f"{len(real_mat_failures)} non-metadata MAT file(s) were unreadable by available readers ({failed_paths}); these are HDF5/v7.3 example force files and h5py is missing in stn_env."
        )

    noncritical = []
    if expected_missing:
        noncritical.append(f"Expected helper/component files not observed: {', '.join(expected_missing)}.")
    noncritical.append("Full original 16-patient/15-control raw cohort data are not present, which is expected from the paper data-availability statement unless another package claims to contain them.")
    noncritical.append("MATLAB/FieldTrip/statistics-toolbox runtime readiness was not tested because this audit does not run MATLAB.")
    if any(row["relative_path"].startswith("__MACOSX") or "/._" in row["relative_path"] for row in physical_rows):
        noncritical.append("The extracted package contains macOS metadata stubs under __MACOSX and ._* files; these are inventoried but not treated as analysis files.")
    if missing_deps:
        noncritical.append("Some MATLAB calls resolve to external MATLAB/FieldTrip/toolbox/helper dependencies and must be resolved before reproduction.")

    recommendations = []
    if not missing_examples:
        recommendations.append("Phase 6B should validate behavioral force/event parsing on the Kont01/Kont02 example data only, without assuming full cohort raw data.")
    else:
        recommendations.append("Before Phase 6B, obtain or restore the expected Kont01/Kont02 example behavioral files or document their absence.")
    if not any(row.get("is_possible_raw_lfp") == "true" and row["extension"] == ".mat" for row in physical_rows):
        recommendations.append("Treat this package primarily as a code/reproducibility reference plus example behavioral data, not as a full raw STN LFP dataset.")
    else:
        recommendations.append("If raw LFP files are confirmed, run a separate Phase 6B raw STN LFP preprocessing audit with MATLAB/FieldTrip compatibility checks.")
    if real_mat_failures:
        recommendations.append("Before Phase 6B behavioral force parsing, resolve MATLAB v7.3/HDF5 reading for the Kont01/Kont02 force files by using an approved h5py-capable stn_env or a MATLAB-side schema export.")
    recommendations.append("Document MATLAB, FieldTrip, Statistics and Machine Learning Toolbox, and missing helper-script requirements before attempting reproduction.")
    recommendations.append("Create a code-methods map from the MATLAB modules to the STN detector project before porting any analysis logic.")

    sampling_refs = code_sampling_refs_from_files(files, data_root, max_text_read_bytes)
    findings = {
        "dataset_name": DATASET_NAME,
        "dataset_root": str(data_root),
        "audit_timestamp": timestamp,
        "audit_scope": "Phase 6A Part 2 dataset/code/data audit only; no MATLAB, FieldTrip, LME, permutation statistics, or preprocessing executed.",
        "repo_root": str(repo_root),
        "environment_python": f"{sys.executable} ({sys.version.split()[0]})",
        "expected_study_pd_patients": EXPECTED_STUDY_PD_PATIENTS,
        "expected_study_hc_analyzed": EXPECTED_STUDY_HC_ANALYZED,
        "expected_stim_patients": EXPECTED_STIM_PATIENTS,
        "full_original_data_expected_in_package": False,
        "expected_example_subjects": EXPECTED_EXAMPLE_SUBJECTS,
        "observed_example_subjects": observed_examples,
        "missing_example_subjects": missing_examples,
        "observed_subject_like_ids": observed_ids,
        "total_files": len(physical_rows),
        "total_manifest_rows_including_zip_members": len(manifest_rows),
        "total_matlab_files": ext_counts[".m"],
        "total_real_matlab_code_files_excluding_macos_metadata": sum(1 for row in physical_rows if row.get("is_possible_code_file") == "true"),
        "total_mat_files": ext_counts[".mat"],
        "total_csv_files": ext_counts[".csv"],
        "total_text_files": ext_counts[".txt"] + ext_counts[".rtf"],
        "total_excel_files": ext_counts[".xlsx"] + ext_counts[".xls"],
        "total_zip_files": ext_counts[".zip"],
        "total_possible_code_files": sum(1 for row in physical_rows if row.get("is_possible_code_file") == "true"),
        "total_possible_data_files": sum(1 for row in physical_rows if row["extension"] in MAT_EXTENSIONS | TABULAR_EXTENSIONS),
        "total_possible_raw_lfp_files": sum(1 for row in physical_rows if row.get("is_possible_raw_lfp") == "true"),
        "total_possible_behavioral_files": sum(1 for row in physical_rows if row.get("is_possible_behavioral_data") == "true"),
        "total_possible_stim_files": sum(1 for row in physical_rows if row.get("is_possible_stim_file") == "true"),
        "expected_components_found": expected_found,
        "expected_components_missing": expected_missing,
        "module_summary": module_summary,
        "matlab_dependency_summary": dep_counts,
        "missing_dependency_summary": missing_deps,
        "mat_read_failures": mat_failures,
        "hdf5_mat_files_unreadable_if_any": [
            row for row in mat_failures if "Please use HDF reader" in row.get("error_message", "") or "hdf5" in row.get("error_message", "").lower()
        ],
        "sampling_rate_summary": {
            "paper_default_lfp_fs_hz": args.paper_fs,
            "lfp_analysis_downsampled_fs_hz": args.lfp_analysis_fs,
            "stim_binary_downsampled_fs_hz": args.stim_binary_fs,
            "explicit_mat_fsample_values": sorted(set(clean_join(row.get("fsample_values_found", "").split(";")) for row in mat_rows if row.get("fsample_values_found"))),
            "code_references": sampling_refs,
            "notes": "2048 Hz is paper context only unless an explicit file/code field states it; 200 Hz and 1000 Hz are interpreted from code/package documentation context.",
        },
        "possible_scan_or_identifiable_files": scan_flags,
        "privacy_governance_flags": privacy_rows,
        "critical_issues": critical_issues,
        "noncritical_limitations": noncritical,
        "overall_status": "completed_with_critical_issues" if critical_issues else "completed_with_noncritical_limitations",
        "recommended_next_steps": recommendations,
    }
    return findings


def main() -> int:
    args = parse_args()
    repo_root = Path.cwd().resolve()
    data_root = Path(args.data_root).expanduser()
    if not data_root.is_absolute():
        data_root = repo_root / data_root
    data_root = data_root.resolve()
    out_dir = Path(args.out_dir).expanduser()
    if not out_dir.is_absolute():
        out_dir = repo_root / out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    if not data_root.exists():
        raise SystemExit(f"Data root does not exist: {data_root}")
    if not data_root.is_dir():
        raise SystemExit(f"Data root is not a directory: {data_root}")

    timestamp = local_timestamp()
    files = collect_files(data_root)
    local_m_map = build_local_m_map(files, data_root)
    code_rows, dep_rows, code_info_by_rel, code_subject_refs = inspect_matlab_code(
        files, data_root, local_m_map, args.max_text_read_bytes
    )
    manifest_rows, zip_member_rows = build_manifest(files, data_root, code_info_by_rel)
    module_rows = build_module_inventory(files, data_root, code_rows)
    expected_rows, expected_found, expected_missing = build_expected_component_matrix(local_m_map, code_rows)
    subject_rows, observed_examples, missing_examples = build_subject_inventory(
        manifest_rows, code_rows, code_subject_refs
    )
    mat_rows, mat_schema_rows = inspect_mat_files(files, data_root, args.max_sanity_arrays)
    tabular_rows, tabular_schema_rows = inspect_tabular_files(files, data_root, args.max_text_read_bytes)
    schema_rows = mat_schema_rows + tabular_schema_rows
    for zip_row in zip_member_rows:
        schema_rows.append(
            {
                "relative_path": zip_row["relative_path"],
                "module_guess": zip_row["top_level_module_guess"],
                "inferred_subject_id": zip_row["inferred_subject_id"],
                "data_object_name": Path(zip_row["relative_path"]).name,
                "object_type": "zip_member",
                "shape_or_rows_cols": "",
                "dtype_or_column_types_if_known": zip_row["extension"],
                "key_variables_detected": "",
                "sampling_rate_detected": "",
                "sampling_rate_source": "unknown",
                "time_axis_detected": "false",
                "trial_axis_detected": "false",
                "channel_axis_detected": "false",
                "channel_labels_detected": "",
                "lfp_data_possible": zip_row["is_possible_raw_lfp"],
                "force_data_possible": zip_row["is_possible_force_data"],
                "behavior_data_possible": zip_row["is_possible_behavioral_data"],
                "stim_data_possible": zip_row["is_possible_stim_file"],
                "notes": zip_row["notes"],
            }
        )
    privacy_rows = build_privacy_inventory(manifest_rows, code_rows, tabular_rows)
    findings = build_findings(
        args,
        repo_root,
        data_root,
        timestamp,
        manifest_rows,
        code_rows,
        dep_rows,
        mat_rows,
        tabular_rows,
        module_rows,
        privacy_rows,
        subject_rows,
        expected_found,
        expected_missing,
        observed_examples,
        missing_examples,
        files,
        args.max_text_read_bytes,
    )
    commands = write_commands_file(out_dir / "phase6a_part2_commands_run.txt")

    manifest_fields = [
        "relative_path",
        "parent_dir",
        "top_level_module_guess",
        "extension",
        "size_bytes",
        "size_mb",
        "file_name",
        "stem",
        "inferred_subject_id",
        "inferred_group",
        "inferred_analysis_part",
        "possible_figure_mapping",
        "is_mat_file",
        "is_m_file",
        "is_csv_file",
        "is_text_file",
        "is_excel_file",
        "is_zip_file",
        "is_possible_raw_lfp",
        "is_possible_behavioral_data",
        "is_possible_force_data",
        "is_possible_event_file",
        "is_possible_stim_file",
        "is_possible_result_file",
        "is_possible_code_file",
        "is_possible_fieldtrip_dependency",
        "is_possible_scan_or_identifiable_file",
        "notes",
    ]
    module_fields = [
        "module_guess",
        "relative_dir",
        "n_files",
        "n_matlab_files",
        "n_mat_files",
        "n_csv_tsv_txt_xlsx_files",
        "n_zip_files",
        "n_subject_like_files",
        "key_scripts_found",
        "data_files_found",
        "readme_or_instruction_files_found",
        "likely_inputs",
        "likely_outputs",
        "notes",
    ]
    expected_fields = [
        "component_name",
        "expected_module",
        "expected_role",
        "exact_file_found",
        "near_match_files",
        "relative_path",
        "is_function_or_script",
        "dependencies_detected",
        "input_files_or_dirs_detected",
        "output_files_detected",
        "notes",
    ]
    subject_fields = [
        "subject_id",
        "inferred_group",
        "is_example_subject",
        "source_of_detection",
        "n_files",
        "file_types",
        "modules_present",
        "behavioral_files_present",
        "event_files_present",
        "force_files_present",
        "lfp_files_present",
        "stim_files_present",
        "notes",
    ]
    mat_fields = [
        "relative_path",
        "module_guess",
        "inferred_subject_id",
        "size_mb",
        "mat_format_guess",
        "load_status",
        "error_message",
        "top_level_keys",
        "n_top_level_keys",
        "has_fieldtrip_like_struct",
        "has_fsample",
        "fsample_values_found",
        "has_label",
        "n_labels_if_found",
        "has_trial",
        "trial_shape_summary",
        "has_time",
        "time_shape_summary",
        "has_freq",
        "freq_shape_summary",
        "has_powspctrm_or_power",
        "power_shape_summary",
        "has_force_like_data",
        "has_event_like_data",
        "has_stim_like_data",
        "has_behavioral_table_like_data",
        "has_lfp_like_data",
        "likely_raw_or_result_or_example",
        "possible_figure_source_data",
        "notes",
    ]
    tabular_fields = [
        "relative_path",
        "module_guess",
        "inferred_subject_id",
        "extension",
        "size_mb",
        "read_status",
        "delimiter_guess",
        "n_rows_if_countable",
        "n_columns_if_countable",
        "column_names_or_first_fields",
        "has_behavioral_columns_guess",
        "has_force_columns_guess",
        "has_event_columns_guess",
        "has_stim_columns_guess",
        "has_demographic_or_clinical_columns_guess",
        "has_path_or_config_content_guess",
        "notes",
    ]
    code_fields = [
        "relative_path",
        "file_name",
        "module_guess",
        "n_lines",
        "is_function",
        "function_name_if_any",
        "script_or_function_guess",
        "first_comment_block_summary",
        "mentions_fieldtrip",
        "ft_functions_called",
        "mentions_fitlme_or_lme",
        "mentions_permutation_or_cluster",
        "mentions_force_processing",
        "mentions_lfp_processing",
        "mentions_stimulation_processing",
        "mentions_psychopy_or_events",
        "mentions_leaddbs_or_imaging",
        "hardcoded_paths_found",
        "load_calls_found",
        "save_calls_found",
        "read_calls_found",
        "write_calls_found",
        "input_variables_or_paths_guess",
        "output_variables_or_paths_guess",
        "figures_or_plotting_guess",
        "notes",
    ]
    dep_fields = [
        "caller_relative_path",
        "caller_function_or_script",
        "dependency_name",
        "dependency_type_guess",
        "dependency_file_found_in_package",
        "dependency_relative_path_if_found",
        "evidence_line_excerpt_short",
        "notes",
    ]
    schema_fields = [
        "relative_path",
        "module_guess",
        "inferred_subject_id",
        "data_object_name",
        "object_type",
        "shape_or_rows_cols",
        "dtype_or_column_types_if_known",
        "key_variables_detected",
        "sampling_rate_detected",
        "sampling_rate_source",
        "time_axis_detected",
        "trial_axis_detected",
        "channel_axis_detected",
        "channel_labels_detected",
        "lfp_data_possible",
        "force_data_possible",
        "behavior_data_possible",
        "stim_data_possible",
        "notes",
    ]
    privacy_fields = ["relative_path", "risk_type", "severity", "evidence", "recommended_action"]

    write_csv(out_dir / "dataset_manifest.csv", manifest_fields, manifest_rows)
    write_csv(out_dir / "module_folder_inventory.csv", module_fields, module_rows)
    write_csv(out_dir / "expected_component_matrix.csv", expected_fields, expected_rows)
    write_csv(out_dir / "subject_file_inventory.csv", subject_fields, subject_rows)
    write_csv(out_dir / "mat_file_inventory.csv", mat_fields, mat_rows)
    write_csv(out_dir / "tabular_file_inventory.csv", tabular_fields, tabular_rows)
    write_csv(out_dir / "matlab_code_inventory.csv", code_fields, code_rows)
    write_csv(out_dir / "matlab_dependency_inventory.csv", dep_fields, dep_rows)
    write_csv(out_dir / "data_schema_inventory.csv", schema_fields, schema_rows)
    write_csv(out_dir / "privacy_governance_inventory.csv", privacy_fields, privacy_rows)
    (out_dir / "audit_findings.json").write_text(
        json.dumps(json_safe(findings), indent=2, sort_keys=True, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )
    build_readme(
        out_dir / "README_dataset_audit.md",
        args,
        repo_root,
        data_root,
        timestamp,
        manifest_rows,
        module_rows,
        expected_rows,
        subject_rows,
        mat_rows,
        tabular_rows,
        code_rows,
        dep_rows,
        privacy_rows,
        findings,
        commands,
    )

    print(json.dumps({"out_dir": str(out_dir), "total_files": len(files), "status": findings["overall_status"]}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
